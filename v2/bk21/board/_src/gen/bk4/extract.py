# -*- coding: utf-8 -*-
# 하는 일: BK4 학교별 명단 엑셀 10개를 읽어 roster_raw.json 으로 만든다
"""8개 대학 참여교수·신진 명단을 한 형태로 정제한다.
   실명은 여기(로컬 작업파일)에만 둔다. 배포용에는 마스킹해서 나간다."""
import openpyxl, os, json, re, unicodedata
# 연세대·한양대는 성명 칸이 비고 마지막 열에 «name: 홍길동 / position: 교수» 덩어리로 들어 있다.
BLOB=re.compile(r'name:\s*([^\n]+)')
POSB=re.compile(r'position:\s*([^\n]+)')
SRC=os.path.expanduser('~/Downloads/BK4 학교 별')
OUT='/private/tmp/claude-501/-Users-karis-dev-biblo-rims-front/9d99337a-c669-4c91-b528-cfc7c84b97fe/scratchpad/bk4'
UNIV={'SNU':'서울대','YS':'연세대','KU':'고려대','SKKU':'성균관대','HYU':'한양대',
      'KHU':'경희대','CAU':'중앙대','SG':'서강대','HUFS':'한국외대'}

def norm(s):
    return unicodedata.normalize('NFC', str(s)).strip() if s is not None else ''

def hdrmap(ws):
    """헤더 행을 찾아 {정규화이름: 열index}"""
    for i,r in enumerate(ws.iter_rows(min_row=1,max_row=5,values_only=True)):
        v=[norm(c) for c in r]
        if '성명' in v:
            m={}
            for j,h in enumerate(v):
                if not h: continue
                k=h.replace(' ','')
                if k=='성명': m['name']=j
                elif k in ('영문성명','영문명'): m['en']=j
                elif k in ('학과','교육연구단(팀)','대학원학과'): m.setdefault('dept',j)
                elif k in ('직위','직위구분'): m['pos']=j
                elif k=='소속': m['aff']=j
                elif k in ('전공분야',): m['major']=j
                elif k=='이메일': m['mail']=j
            return i+1, m
    return None, {}

rows=[]
for f in sorted(os.listdir(SRC)):
    if not f.endswith('.xlsx'): continue
    code=norm(f).split('_')[0]
    if code not in UNIV: continue
    wb=openpyxl.load_workbook(os.path.join(SRC,f), read_only=True)
    for kind,key in (('참여교수','prof'),('신진연구인력','new')):
        ws=next((x for x in wb.worksheets if kind in x.title), None)
        if not ws: continue
        hr,m=hdrmap(ws)
        if not m.get('name'): continue
        for r in ws.iter_rows(min_row=hr+1, values_only=True):
            g=lambda k: norm(r[m[k]]) if k in m and m[k]<len(r) else ''
            nm=g('name'); pos=g('pos')
            if not nm:                                  # 덩어리 열에서 회수
                for c in reversed(r):
                    t=norm(c)
                    if 'name:' in t:
                        mm=BLOB.search(t)
                        if mm: nm=mm.group(1).strip()
                        pp=POSB.search(t)
                        if pp and not pos: pos=pp.group(1).strip()
                        break
            nm=re.sub(r'[（(].*$','',nm).strip()          # 「주영준(Zhu, Yongjun)」 → 「주영준」
            if not nm or len(nm)>12 or nm in ('성명',): continue
            rows.append({'u':code, 'un':UNIV[code], 'kind':key,
                'name':nm, 'en':g('en'), 'dept':g('dept'), 'pos':pos,
                'aff':g('aff'), 'major':g('major')[:60], 'mail':g('mail')})
    wb.close()

json.dump(rows, open(f"{OUT}/roster_raw.json","w"), ensure_ascii=False)
print(f"  총 {len(rows):,}명")
from collections import Counter
c=Counter((x['u'],x['kind']) for x in rows)
for u in ['SNU','YS','KU','SKKU','HYU','KHU','CAU','SG']:
    print(f"  {UNIV[u]:6s} 참여교수 {c[(u,'prof')]:5d} · 신진 {c[(u,'new')]:4d}")
en=sum(1 for x in rows if x['en'])
print(f"\n  영문성명 보유 {en:,}/{len(rows):,} = {en/len(rows)*100:.1f}%")
d=Counter(x['dept'] for x in rows if x['dept'])
print(f"  학과 고유 {len(d)}종 · 상위 8: {[k for k,_ in d.most_common(8)]}")
